#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
Topic: 从数据库中过滤出广州的企业，然后导出为excel格式
Desc :
"""

import sys
import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.drawing import Image
from openpyxl.writer.dump_worksheet import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors, borders, fills

import logging
import datetime
import mysql.connector
from mysql.connector import errorcode

customer_type_map = {
    7: '事务所',
    5: '其他',
    1: '个人',
    2: '税务机关',
    3: '企业',
    4: '经销商',
    6: '集团客户',
    8: '公安'
}
enterprise_type_map = {
    1: '国有企业',
    2: '集体企业',
    3: '股份合作企业',
    4: '联营企业',
    5: '有限责任公司',
    6: '股份有限公司',
    7: '私营企业',
    8: '其他企业',
    9: '合资经营企业（港或澳、台资）',
    10: '合作经营企业（港或澳、台资）',
    11: '港、澳、台商独资经营企业',
    12: '港、澳、台商投资股份有限公司',
    13: '中外合资经营企业',
    14: '中外合作经营企业',
    15: '外资企业',
    16: '外商投资股份有限公司',
    17: '个体工商户'
}

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
_log = logging.getLogger('app.' + __name__)

SQL_ID_PARENTID = """
DELIMITER $$
CREATE FUNCTION getChildLst (rootId BIGINT)
	RETURNS VARCHAR (1000) DETERMINISTIC
	BEGIN
		DECLARE sTemp VARCHAR (1000) ;
		DECLARE sTempChd VARCHAR (1000) ;
		SET sTemp = '$' ;
		SET sTempChd = cast(rootId AS CHAR) ;
		WHILE sTempChd IS NOT NULL DO
			SET sTemp = concat(sTemp, ',', sTempChd) ;
			SELECT group_concat(id) INTO sTempChd
			FROM
				t_region
			WHERE
				FIND_IN_SET(parent_id, sTempChd) > 0 ;
		END WHILE ;
		RETURN sTemp ;
	END$$
DELIMITER ;
"""

SQL_SELECT_REGION_ID = """
SELECT getChildLst(%s) as rid;
"""

SQL_SELECT_GZ = """
SELECT
    A.id AS id,
    A.name AS name,
    A.tax_code AS tax_code,
    A.region_id AS region_id,
    A.customer_type AS customer_type,
    A.enterprise_type AS enterprise_type,
    A.address AS address,
    A.postcode AS postcode,
    A.tel AS tel,
    A.contact AS contact,
    A.fax AS fax,
    A.mobile AS mobile,
    B.region_code as region_code,
    B.regian_name as regian_name,
    B.note as note,
    B.parent_id as parent_id
FROM t_enterprise A
LEFT OUTER JOIN t_region B ON A.region_id=B.id
WHERE B.id IN ({})
"""


def _connect():
    config = {
        'user': 'root',
        'password': 'mysql',
        'host': '192.168.203.95',
        'database': 'hangxin',
        'raise_on_warnings': True,
    }
    cnx = None
    try:
        cnx = mysql.connector.connect(**config)
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(err)
        if cnx:
            cnx.close()
    return cnx


def load_gz_data():
    _log.info('开始从数据库中加载数据')
    conn = _connect()
    cur = conn.cursor()
    # 广州的region_id为2152
    cur.execute(SQL_SELECT_REGION_ID, (2152,))
    rids = cur.fetchall()[0][0]
    print('rids----{}'.format(rids))
    cur.execute(SQL_SELECT_GZ.format(rids[2:]))
    result = [list(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    for v in result:
        if v[4]:
            v[4] = customer_type_map[int(v[4])]
        if v[5]:
            v[5] = enterprise_type_map[int(v[5])]
    _log.info('数据库中加载数据完毕')
    return result


def export_to_excel(db_data, xlsx_name):
    """导出到excel文件中"""
    _log.info('开始导出到excel文件中')
    border = Border(
        left=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        right=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')
    )
    alignment = Alignment(horizontal='justify',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=True,
                          indent=0)
    fill = PatternFill(fill_type=None, start_color='FFFFFFFF')
    # 基本的样式
    basic_style = Style(font=Font(name='Microsoft YaHei')
                        , border=border, alignment=alignment
                        , fill=fill)
    header_style = basic_style.copy(
        font=Font(name='Microsoft YaHei', b=True, size=15, color='00215757'),
        fill=PatternFill(fill_type=fills.FILL_SOLID, start_color='00BAA87F'))
    common_style = basic_style.copy()
    wb = Workbook()
    ws = wb.create_sheet(index=0, title='enterprises-{}'.format(len(db_data)))

    ws['A1'] = 'id'
    ws['A1'].style = common_style
    ws['B1'] = 'name'
    ws['B1'].style = common_style
    ws['C1'] = 'tax_code'
    ws['C1'].style = common_style
    ws['D1'] = 'region_id'
    ws['D1'].style = common_style
    ws['E1'] = 'customer_type'
    ws['E1'].style = common_style
    ws['F1'] = 'enterprise_type'
    ws['F1'].style = common_style
    ws['G1'] = 'address'
    ws['G1'].style = common_style
    ws['H1'] = 'postcode'
    ws['H1'].style = common_style
    ws['I1'] = 'tel'
    ws['I1'].style = common_style
    ws['J1'] = 'contact'
    ws['J1'].style = common_style
    ws['K1'] = 'fax'
    ws['K1'].style = common_style
    ws['L1'] = 'mobile'
    ws['L1'].style = common_style
    ws['M1'] = 'region_code'
    ws['M1'].style = common_style
    ws['N1'] = 'regian_name'
    ws['N1'].style = common_style
    ws['O1'] = 'note'
    ws['O1'].style = common_style
    ws['P1'] = 'parent_id'
    ws['P1'].style = common_style

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 80
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20

    for i, row in enumerate(db_data):
        ws['A{}'.format(i + 2)] = row[0]
        ws['A{}'.format(i + 2)].style = common_style
        ws['B{}'.format(i + 2)] = row[1]
        ws['B{}'.format(i + 2)].style = common_style
        ws['C{}'.format(i + 2)] = row[2]
        ws['C{}'.format(i + 2)].style = common_style
        ws['D{}'.format(i + 2)] = row[3]
        ws['D{}'.format(i + 2)].style = common_style
        ws['E{}'.format(i + 2)] = row[4]
        ws['E{}'.format(i + 2)].style = common_style
        ws['F{}'.format(i + 2)] = row[5]
        ws['F{}'.format(i + 2)].style = common_style
        ws['G{}'.format(i + 2)] = row[6]
        ws['G{}'.format(i + 2)].style = common_style
        ws['H{}'.format(i + 2)] = row[7]
        ws['H{}'.format(i + 2)].style = common_style
        ws['I{}'.format(i + 2)] = row[8]
        ws['I{}'.format(i + 2)].style = common_style
        ws['J{}'.format(i + 2)] = row[9]
        ws['J{}'.format(i + 2)].style = common_style
        ws['K{}'.format(i + 2)] = row[10]
        ws['K{}'.format(i + 2)].style = common_style
        ws['L{}'.format(i + 2)] = row[11]
        ws['L{}'.format(i + 2)].style = common_style
        ws['M{}'.format(i + 2)] = row[12]
        ws['M{}'.format(i + 2)].style = common_style
        ws['N{}'.format(i + 2)] = row[13]
        ws['N{}'.format(i + 2)].style = common_style
        ws['O{}'.format(i + 2)] = row[14]
        ws['O{}'.format(i + 2)].style = common_style
        ws['P{}'.format(i + 2)] = row[15]
        ws['P{}'.format(i + 2)].style = common_style
    wb.save(filename=xlsx_name)
    _log.info('导出excel文件完成')


if __name__ == '__main__':
    gz_data = load_gz_data()
    export_to_excel(gz_data, r'D:\download\20150505\gz_enterprises.xlsx')
    pass

