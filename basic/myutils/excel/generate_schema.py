#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
Topic: 通过一个schema.sql来生成excel表格的数据库设计文档
Desc : 
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.drawing import Image
from openpyxl.writer.dump_worksheet import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles import colors, borders, fills


def load_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "首页列表"
    ws = wb['首页列表']
    print(wb.get_sheet_names())
    print(ws['D5'], ws.cell(row=5, column=4))
    cell_range = ws['A1':'C2']

    wb2 = load_workbook('D:/work/MySQL数据库表.xlsx')
    print(wb2.get_sheet_names())


def write_xlsx():
    wb = Workbook()
    dest_filename = 'empty_book.xlsx'
    ws = wb.active
    ws.title = "首页列表"
    for col_idx in range(1, 10):
        col = get_column_letter(col_idx)
        for row in range(1, 20):
            ws['%s%s' % (col, row)].value = '%s%s' % (col, row)
    ws.merge_cells('A1:B1')  # 合并单元格
    ws.unmerge_cells('A1:B1')
    ws = wb.create_sheet()
    ws.title = 'Pi'
    ws['F5'] = 3.14
    # img = Image('logo.png')
    # img.drawing.top = 100
    # img.drawing.left = 150

    wb.save(filename=dest_filename)

    wb = load_workbook(filename='empty_book.xlsx')
    sheet_ranges = wb['首页列表']
    print(sheet_ranges['D18'].value)


def write_only():
    wb = Workbook()
    ws = wb.create_sheet()
    ws.title = "首页列表"
    c = ws['A1']
    c.style = Style(font=Font(name='Courrier', size=36)
                    , fill=PatternFill(fill_type=None, start_color='FFFFFFFF',
                                       end_color='FF000000')
                    , protection=Protection(locked='inherit', hidden='inherit')
                    , alignment=Alignment(horizontal='general', vertical='bottom',
                                          shrink_to_fit=True)
                    , border=Border(left=Side(border_style=None, color='FF000000')))
    c.value = '姓名'
    # cell = WriteOnlyCell(ws, value="hello world")
    # cell.style = Style(font=Font(name='Courrier', size=36))
    # cell.comment = Comment(text="A comment", author="Author's Name")

    # ws.header_footer.center_header.text = 'My Excel Page'
    # ws.header_footer.center_header.font_size = 14
    # ws.header_footer.center_header.font_name = "Tahoma,Bold"
    # ws.header_footer.center_header.font_color = "CC3366"
    wb.save(filename='empty_book.xlsx')


def load_schema(filename):
    """先加载schema.sql文件来获取所有建表语句"""
    result = []
    with open(filename, encoding='utf-8') as sqlfile:
        each_table = []  # 每张表定义
        for line in sqlfile:
            if line.startswith('--'):
                each_table.insert(0, line.split('--')[1].strip())
            elif 'DROP TABLE' in line:
                each_table.insert(1, line.strip().split()[-1][:-1])
            elif ' COMMENT ' in line and 'ENGINE=' not in line:
                col_arr = line.split()
                col_name = col_arr[0]
                col_type = col_arr[1]
                if 'PRIMARY KEY' in line or 'NOT NULL' in line:
                    col_null = 'NOT NULL'
                else:
                    col_null = ''
                col_remark = col_arr[-1]
                each_table.append((col_name, col_type, col_null, col_remark))
            elif 'ENGINE=' in line:
                # 单个表定义结束
                result.append(list(each_table))
                each_table.clear()
    return result


def write_dest(xlsx_name, schema_name):
    border = Border(
        left=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        right=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')
    )
    alignment = Alignment(horizontal='justify', vertical='bottom',
                          text_rotation=0, wrap_text=False,
                          shrink_to_fit=True, indent=0)
    fill = PatternFill(fill_type=None, start_color='FFFFFFFF')
    # 基本的样式
    basic_style = Style(font=Font(name='Microsoft YaHei')
                        , border=border, alignment=alignment
                        , fill=fill)
    title_style = basic_style.copy(
        font=Font(name='Microsoft YaHei', b=True, size=20, color='00215757'),
        alignment=Alignment(horizontal='center', vertical='bottom',
                            text_rotation=0, wrap_text=False,
                            shrink_to_fit=True, indent=0),
        fill=PatternFill(fill_type=fills.FILL_SOLID, start_color='00B2CBED'))
    header_style = basic_style.copy(
        font=Font(name='Microsoft YaHei', b=True, size=15, color='00215757'),
        fill=PatternFill(fill_type=fills.FILL_SOLID, start_color='00BAA87F'))
    common_style = basic_style.copy()
    link_style = basic_style.copy(font=Font(
        name='Microsoft YaHei', color=colors.BLUE, underline='single'))
    table_data = load_schema(schema_name)
    wb = Workbook()
    wb.active.title = "首页列表"

    for table in table_data:
        ws = wb.create_sheet(title=table[0])
        ws.merge_cells('E3:H3')  # 合并单元格
        ws['E3'].style = title_style
        ws['F2'].style = Style(border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['G2'].style = Style(border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['H2'].style = Style(border=Border(
            bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['I3'].style = Style(border=Border(
            left=Side(border_style=borders.BORDER_THIN, color='FF000000')))
        ws['E3'] = table[0]
        ws['E4'].style = header_style
        ws['E4'] = '列名'
        ws['F4'].style = header_style
        ws['F4'] = '类型'
        ws['G4'].style = header_style
        ws['G4'] = '空值约束'
        ws['H4'].style = header_style
        ws['H4'] = '备注'
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 45
        for idx, each_column in enumerate(table[2:]):
            ws['E{}'.format(idx + 5)].style = common_style
            ws['E{}'.format(idx + 5)] = each_column[0]
            ws['F{}'.format(idx + 5)].style = common_style
            ws['F{}'.format(idx + 5)] = each_column[1]
            ws['G{}'.format(idx + 5)].style = common_style
            ws['G{}'.format(idx + 5)] = each_column[2]
            ws['H{}'.format(idx + 5)].style = common_style
            ws['H{}'.format(idx + 5)] = each_column[3].strip().split('\'')[1]
    ws = wb['首页列表']
    ws.merge_cells('D3:F3')
    ws['D3'].style = title_style
    ws['E2'].style = Style(border=Border(
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
    ws['F2'].style = Style(border=Border(
        bottom=Side(border_style=borders.BORDER_THIN, color='FF000000')))
    ws['G3'].style = Style(border=Border(
        left=Side(border_style=borders.BORDER_THIN, color='FF000000')))
    ws['D3'] = '贷快发数据库系统表'
    ws['D4'].style = header_style
    ws['D4'] = '编号'
    ws['E4'].style = header_style
    ws['E4'] = '表名'
    ws['F4'].style = header_style
    ws['F4'] = '详情链接'
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    for inx, val in enumerate(table_data):
        ws['D{}'.format(inx + 5)].style = common_style
        ws['D{}'.format(inx + 5)] = inx + 1
        ws['E{}'.format(inx + 5)].style = common_style
        ws['E{}'.format(inx + 5)] = val[1]
        linkcell = ws['F{}'.format(inx + 5)]
        linkcell.style = link_style
        linkcell.value = val[0]
        linkcell.hyperlink = '#{0}!{1}'.format(val[0], 'E3')
    wb.save(filename=xlsx_name)


if __name__ == '__main__':
    # write_xlsx()
    # write_only()
    write_dest('贷快发数据库设计.xlsx', 'schema.sql')
    pass
